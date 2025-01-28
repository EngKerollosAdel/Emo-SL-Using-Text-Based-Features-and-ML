import openpyxl


class Configuration:
    # Public strings
    class FileName:
        stopWordsFileName = "Data/stop_words.txt"
        positive_lexicon = "Data/positive_lexicon.txt"
        negative_lexicon = "Data/negative_lexicon.txt"
        emojis = "Data/emojis.txt"
 
    # Public strings
    class SheetName:
        sentiment_lexicon = "sentiment_lexicon"
        emoji_sentiment_lexicon = "EMOJI_SENTIMENT_LEXICON"

    # Public strings
    class Sentiment:
        positive = "positive"
        negative = "negative"

    @staticmethod
    def fetch_tweets():
        """
        Fetch tweets from a text file.
        """
        tweets = []
        try:
            with open('Data/tweets.txt', 'r', encoding='utf-8') as file:
                for line in file:
                    tweet = line.strip()
                    if tweet:
                        tweets.append(tweet)
        except FileNotFoundError:
            print(f"The file 'Data/tweets.txt' was not found.")
        return tweets
    

    def fetch_data_from_file(fileName ):
        """
        Fetch tweets from a text file.
        """
        tweets = []
        try:
            with open(fileName, 'r', encoding='utf-8') as file:
                for line in file:
                    tweet = line.strip()
                    if tweet:
                        tweets.append(tweet)
        except FileNotFoundError:
            print(f"The file the targt file was not found.")
        return tweets

    @staticmethod
    def fetch_setting(sheet_name):
        """
        Fetch the sentiment lexicon from an Excel file.

        :param sheet_name: Name of the sheet to read from.
        :return: A dictionary where keys are words and values are their sentiment scores.
        """
        file_name = 'Data/sentiment_lexicon.xlsx'
        setting = {}

        try:
            # Load the workbook
            workbook = openpyxl.load_workbook(file_name)

            # Select the specified sheet or the first sheet
            sheet = workbook[sheet_name] if sheet_name else workbook.active

            # Iterate through the rows (assuming the first row contains headers)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                word, sentiment = row
                if word and sentiment is not None:
                    setting[word.strip()] = float(sentiment)
        except FileNotFoundError:
            print(f"The file '{file_name}' was not found.")
        except KeyError:
            print(f"The sheet '{sheet_name}' does not exist in the workbook.")
        except Exception as e:
            print(f"An error occurred: {e}")
        return setting
    
    @staticmethod
    def fetch_setting_pattern(sheet_name):
        """
        Fetch the sentiment lexicon from an Excel file.

        :param sheet_name: Name of the sheet to read from.
        :return: A dictionary where keys are words and values are their sentiment scores as strings.
        """
        file_name = 'Data/sentiment_lexicon.xlsx'
        setting = {}

        try:
            # Load the workbook
            workbook = openpyxl.load_workbook(file_name)

            # Select the specified sheet or the first sheet
            sheet = workbook[sheet_name] if sheet_name else workbook.active

            # Iterate through the rows (assuming the first row contains headers)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                word, sentiment = row
                if word and sentiment is not None:
                    # Store both as strings in the dictionary
                    setting[str(word).strip()] = f"{str(sentiment).strip()}"
        except FileNotFoundError:
            print(f"The file '{file_name}' was not found.")
        except KeyError:
            print(f"The sheet '{sheet_name}' does not exist in the workbook.")
        except Exception as e:
            print(f"An error occurred: {e}")
        return setting




def main():
    # Testing the configuration class to fetch tweets from the text file
    tweets = Configuration.fetch_tweets()  # Ensure the file 'tweets.txt' is in the correct folder
    print("Fetched Tweets:")
    for tweet in tweets:
        print(tweet)

    # Testing the fetch_setting method to fetch sentiment lexicon
    sentiment_lexicon = Configuration.fetch_setting('sentiment_lexicon')  # Replace 'sentiment_lexicon' with your sheet name
    print("\nFetched Sentiment Lexicon:")
    print(sentiment_lexicon)

    # Testing the fetch_setting method to fetch sentiment lexicon
    sentiment_lexicon = Configuration.fetch_setting('EMOJI_SENTIMENT_LEXICON')  # Replace 'sentiment_lexicon' with your sheet name
    print("\nFetched EMOJI SENTIMENT LEXICON:")
    print(sentiment_lexicon)


    # Testing the fetch_setting method to fetch sentiment lexicon
    sentiment_lexicon = Configuration.fetch_setting_pattern('normalization_patterns')  # Replace 'sentiment_lexicon' with your sheet name
    print("\nFetched normalization patterns:")
    print(sentiment_lexicon)

if __name__ == "__main__":
    main()
